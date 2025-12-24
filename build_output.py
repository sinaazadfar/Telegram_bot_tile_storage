import argparse
import re
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
import pdfplumber


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u200c", " ")
    text = text.replace("ي", "ی").replace("ك", "ک")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_header(value: object) -> str:
    return normalize_text(value).lower()


_ARABIC_RE = re.compile(r"[\u0600-\u06FF]")


def fix_pdf_text(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    tokens = text.split()
    if not tokens:
        return text
    arabic_tokens = sum(1 for token in tokens if _ARABIC_RE.search(token))
    if arabic_tokens < max(1, len(tokens) // 2):
        return text
    tokens = list(reversed(tokens))
    fixed_tokens: list[str] = []
    for token in tokens:
        if _ARABIC_RE.search(token):
            fixed_tokens.append(token[::-1])
        else:
            fixed_tokens.append(token)
    return " ".join(fixed_tokens)


_CODE_RE = re.compile(r"\d+")
_HEADER_RE = re.compile(r"^(.*?)\s*\((.*?)\)\s*$")
_TONALITY_PREFIX_RE = re.compile(r"^(?:متراژ\s*)?تنالیته\s*(.+)$")

BASE_HEADERS = {
    normalize_header("کد محصول"): "code",
    normalize_header("نام طرح"): "name",
    normalize_header("سایز محصول"): "size",
    normalize_header("عنوان برای محصول"): "size",
    normalize_header("مقدار تقسیم پالت"): "pallet_divisor",
}

METRIC_BASE_HEADERS = {
    normalize_header("A/2"): "a2_meter",
    normalize_header("C/3"): "c3_meter",
    normalize_header("مجموع طرح"): "total_meter",
    normalize_header("متراژ A/2"): "a2_meter",
    normalize_header("متراژ C/3"): "c3_meter",
    normalize_header("مجموع متراژ طرح"): "total_meter",
    normalize_header("متراژ در هر تنالینه"): "per_tonality_avg",
}

METRIC_LABELS = {
    normalize_header("فیزیکی"): "physical",
    normalize_header("قابل فروش"): "sellable",
    normalize_header("رزرو"): "reserved",
    "physical": "physical",
    "sellable": "sellable",
    "reserved": "reserved",
}

METRIC_INPUT_INDEX = {"sellable": 0, "reserved": 1, "physical": 2}
INPUT_HEADERS = {
    normalize_header("موجودی قابل فروش"): "sellable",
    normalize_header("موجودی رزرو"): "reserved",
    normalize_header("موجودی فیزیکی"): "physical",
    normalize_header("نام کالا"): "name",
    normalize_header("نام کاال"): "name",
    normalize_header("درجه"): "degree",
    normalize_header("تنالیته"): "tonality",
    normalize_header("کد کالا"): "code",
    normalize_header("کد کاال"): "code",
}
INPUT_DEFAULT_INDEX = {
    "sellable": 0,
    "reserved": 1,
    "physical": 2,
    "name": 5,
    "degree": 6,
    "tonality": 7,
    "code": 9,
}
LABEL_DISPLAY = {
    "physical": "فیزیکی",
    "sellable": "قابل فروش",
    "reserved": "رزرو",
}
BASE_OUTPUT_HEADERS = ["A/2", "C/3", "مجموع طرح"]


def extract_code_any(value: object) -> int | None:
    if value is None:
        return None
    matches = _CODE_RE.findall(str(value))
    if not matches:
        return None
    try:
        return int(matches[-1])
    except ValueError:
        return None


def code_matches(plan_code: int | None, code_value: object) -> bool:
    if plan_code is None:
        return False
    if code_value is None:
        return False
    plan_text = str(plan_code)
    matches = _CODE_RE.findall(str(code_value))
    if not matches:
        return False
    return any(seq == plan_text for seq in matches if len(seq) == len(plan_text))


def normalize_degree(value: object) -> str:
    text = normalize_text(value)
    if text in {"A/2", "C/3"}:
        return text
    if text.startswith("گرید"):
        if "A" in text:
            return "A/2"
        if "C" in text:
            return "C/3"
    return text


def normalize_tonality(value: object) -> str:
    text = normalize_text(value)
    return text.upper()


def collect_tonalities(rows: list[dict]) -> list[str]:
    seen: set[str] = set()
    tonalities: list[str] = []
    for row in rows:
        tonality = row.get("tonality", "")
        if not tonality or tonality in seen:
            continue
        seen.add(tonality)
        tonalities.append(tonality)
    return tonalities


def as_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def round2(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def format_decimal(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = format(quantized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def format_meter_with_pallets(meter: Decimal, divisor: Decimal) -> str | None:
    if meter <= 0:
        return None
    if divisor <= 0:
        return format_decimal(meter)
    pallets = meter / divisor
    return f"({format_decimal(pallets)}) {format_decimal(meter)}"


def build_header_maps(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[dict, dict, dict]:
    base_cols: dict[str, int] = {}
    metric_cols: dict[str, dict[str, int]] = {}
    tonality_cols: dict[str, dict[str, int]] = {}

    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        header_norm = normalize_header(raw)
        if not header_norm:
            continue
        if header_norm in BASE_HEADERS:
            base_cols[BASE_HEADERS[header_norm]] = col
            continue
        raw_text = normalize_text(raw)
        match = _HEADER_RE.match(raw_text)
        if match:
            base_text = match.group(1)
            label_text = match.group(2)
            base_norm = normalize_header(base_text)
            label_norm = normalize_header(label_text)
            base_key = METRIC_BASE_HEADERS.get(base_norm)
            label_key = METRIC_LABELS.get(label_norm)
            if label_key:
                tonality_match = _TONALITY_PREFIX_RE.match(base_text)
                if tonality_match:
                    tonality_key = normalize_tonality(tonality_match.group(1))
                    if tonality_key:
                        tonality_cols.setdefault(label_key, {})[tonality_key] = col
                    continue
            if base_key and label_key:
                metric_cols.setdefault(label_key, {})[base_key] = col
            continue
        base_key = METRIC_BASE_HEADERS.get(header_norm)
        if base_key:
            metric_cols.setdefault("__default__", {})[base_key] = col

    return base_cols, metric_cols, tonality_cols


def build_input_header_map(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, int]:
    cols: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(1, col).value)
        key = INPUT_HEADERS.get(header)
        if key:
            cols[key] = col
    return cols


def build_input_header_map_from_values(headers: list[object]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for idx, header in enumerate(headers):
        header_norm = normalize_header(header)
        key = INPUT_HEADERS.get(header_norm)
        if key:
            cols[key] = idx
    return cols


def build_input_row(
    name_raw: object,
    code_raw: object,
    degree_raw: object,
    tonality_raw: object,
    sellable_raw: object,
    reserved_raw: object,
    physical_raw: object,
) -> dict:
    sellable = as_decimal(sellable_raw)
    reserved = as_decimal(reserved_raw)
    physical = (
        as_decimal(physical_raw) if physical_raw is not None else sellable + reserved
    )
    name_norm = normalize_text(name_raw)
    return {
        "metrics": {
            "sellable": sellable,
            "reserved": reserved,
            "physical": physical,
        },
        "name_norm": name_norm,
        "degree": normalize_degree(degree_raw),
        "tonality": normalize_tonality(tonality_raw),
        "code_any": extract_code_any(code_raw),
        "code_text": "" if code_raw is None else str(code_raw),
    }


def load_input_rows_xlsx(
    input_path: Path, sheet: str | None
) -> list[dict]:
    in_wb = openpyxl.load_workbook(input_path, data_only=True)
    try:
        in_ws = in_wb[sheet] if sheet else in_wb.active
        input_cols = build_input_header_map(in_ws)

        def col_index(key: str) -> int:
            return input_cols.get(key, INPUT_DEFAULT_INDEX[key]) - 1

        def row_value(row: tuple, key: str) -> object:
            idx = col_index(key)
            if idx < 0 or idx >= len(row):
                return None
            return row[idx]

        input_rows: list[dict] = []
        for row in in_ws.iter_rows(min_row=2, values_only=True):
            name_raw = row_value(row, "name")
            code_raw = row_value(row, "code")
            input_rows.append(
                build_input_row(
                    name_raw,
                    code_raw,
                    row_value(row, "degree"),
                    row_value(row, "tonality"),
                    row_value(row, "sellable"),
                    row_value(row, "reserved"),
                    row_value(row, "physical"),
                )
            )
    finally:
        in_wb.close()
    return input_rows


def load_input_rows_pdf(input_path: Path) -> list[dict]:
    input_rows: list[dict] = []
    with pdfplumber.open(input_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table:
                    continue
                header_row = table[0]
                header_fixed = [fix_pdf_text(cell) for cell in header_row]
                header_map = build_input_header_map_from_values(header_fixed)
                if not header_map:
                    continue
                for row in table[1:]:
                    if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                        continue

                    def cell_value(key: str) -> object:
                        idx = header_map.get(key)
                        if idx is None or idx >= len(row):
                            return None
                        return row[idx]

                    name_raw = fix_pdf_text(cell_value("name"))
                    code_raw = cell_value("code")
                    input_rows.append(
                        build_input_row(
                            name_raw,
                            code_raw,
                            cell_value("degree"),
                            cell_value("tonality"),
                            cell_value("sellable"),
                            cell_value("reserved"),
                            cell_value("physical"),
                        )
                    )
    return input_rows


def fallback_metric_cols() -> dict[str, dict[str, int]]:
    return {
        "__default__": {
            "a2_meter": 5,
            "c3_meter": 6,
            "total_meter": 7,
            "per_tonality_avg": 10,
        }
    }


def ensure_metric_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    tonalities: list[str],
) -> None:
    start_col = ws.max_column + 1
    headers: list[str] = []
    for label_key in ("physical", "sellable", "reserved"):
        label = LABEL_DISPLAY[label_key]
        for base_header in BASE_OUTPUT_HEADERS:
            headers.append(f"{base_header} ({label})")
        for tonality in tonalities:
            headers.append(f"تنالیته {tonality} ({label})")
    for offset, header in enumerate(headers):
        ws.cell(1, start_col + offset).value = header


def summarize_metrics(
    a2_rows: list[dict],
    c3_rows: list[dict],
    metric_key: str,
) -> dict:
    all_rows = a2_rows + c3_rows
    a2_meter = sum((m["metrics"][metric_key] for m in a2_rows), Decimal("0"))
    c3_meter = sum((m["metrics"][metric_key] for m in c3_rows), Decimal("0"))
    total_meter = a2_meter + c3_meter
    per_tonality_map: dict[str, Decimal] = {}
    for row in all_rows:
        tonality = row["tonality"]
        if not tonality:
            continue
        per_tonality_map[tonality] = per_tonality_map.get(tonality, Decimal("0")) + row[
            "metrics"
        ][metric_key]

    total_pallets = len(
        {m["tonality"] for m in all_rows if m["metrics"][metric_key] > 0}
    )

    per_tonality_avg = (
        (total_meter / Decimal(total_pallets)) if total_pallets else Decimal("0")
    )

    return {
        "a2_meter": a2_meter,
        "c3_meter": c3_meter,
        "total_meter": total_meter,
        "per_tonality_avg": per_tonality_avg,
        "per_tonality_map": per_tonality_map,
        "has_pallets": total_pallets > 0,
    }


def write_summary(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    cols: dict[str, int],
    summary: dict,
    tonality_cols: dict[str, int] | None,
    divisor: Decimal,
) -> None:
    if "a2_meter" in cols:
        ws.cell(row, cols["a2_meter"]).value = format_meter_with_pallets(
            summary["a2_meter"], divisor
        )
    if "c3_meter" in cols:
        ws.cell(row, cols["c3_meter"]).value = format_meter_with_pallets(
            summary["c3_meter"], divisor
        )
    if "total_meter" in cols:
        ws.cell(row, cols["total_meter"]).value = format_meter_with_pallets(
            summary["total_meter"], divisor
        )
    if "per_tonality_avg" in cols:
        ws.cell(row, cols["per_tonality_avg"]).value = (
            format_meter_with_pallets(summary["per_tonality_avg"], divisor)
            if summary["has_pallets"]
            else None
        )
    if tonality_cols:
        for tonality, col in tonality_cols.items():
            value = summary["per_tonality_map"].get(tonality, Decimal("0"))
            ws.cell(row, col).value = format_meter_with_pallets(value, divisor)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Fill an output template from input.xlsx inventory (Excel)."
    )
    p.add_argument("--input", default="input.xlsx", help="Path to input.xlsx")
    p.add_argument(
        "--template",
        default="template.xlsx",
        help="Path to template file (will not be modified unless --in-place).",
    )
    p.add_argument(
        "--output",
        default="output_from_template.xlsx",
        help="Path for generated file (ignored when --in-place).",
    )
    p.add_argument(
        "--metric",
        choices=["sellable", "physical", "reserved"],
        default="physical",
        help="Which input column to use for مترمربع aggregation.",
    )
    p.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to use (default: active sheet).",
    )
    p.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite the template file instead of writing --output.",
    )
    return p.parse_args()


def process_files(
    input_path: str | Path,
    template_path: str | Path,
    output_path: str | Path,
    metric: str = "physical",
    sheet: str | None = None,
    in_place: bool = False,
) -> Path:
    if metric not in METRIC_INPUT_INDEX:
        raise ValueError(f"Unsupported metric: {metric}")
    input_path = Path(input_path)
    template_path = Path(template_path)
    output_path = template_path if in_place else Path(output_path)

    if input_path.suffix.lower() == ".pdf":
        input_rows = load_input_rows_pdf(input_path)
    else:
        input_rows = load_input_rows_xlsx(input_path, sheet)

    out_wb = openpyxl.load_workbook(template_path)
    try:
        out_ws = out_wb[sheet] if sheet else out_wb.active

        base_cols, metric_cols, tonality_cols = build_header_maps(out_ws)
        if not metric_cols:
            tonalities = sorted(collect_tonalities(input_rows))
            ensure_metric_headers(out_ws, tonalities)
            base_cols, metric_cols, tonality_cols = build_header_maps(out_ws)
        if not metric_cols:
            metric_cols = fallback_metric_cols()

        name_col = base_cols.get("name", 1)
        code_col = base_cols.get("code", 3)
        divisor_col = base_cols.get("pallet_divisor")

        output_code_counts: Counter[int] = Counter()
        output_codes: set[int] = set()
        for r in range(2, out_ws.max_row + 1):
            code = out_ws.cell(r, code_col).value
            if code is None:
                continue
            try:
                code_int = int(str(code))
            except Exception:
                continue
            output_code_counts[code_int] += 1
            output_codes.add(code_int)

        def find_matches(plan_name: str, plan_code: int | None) -> list[dict]:
            plan_norm = normalize_text(plan_name)
            if not plan_norm:
                return []

            name_matches = [r for r in input_rows if plan_norm in r["name_norm"]]
            if plan_code is not None:
                name_matches = [
                    r
                    for r in name_matches
                    if (not r["code_text"])
                    or code_matches(plan_code, r["code_text"])
                    or (r["code_any"] is None)
                    or (r["code_any"] not in output_codes)
                ]

            if name_matches:
                return name_matches

            if plan_code is None:
                return []
            if output_code_counts[plan_code] > 1:
                return []
            return [r for r in input_rows if code_matches(plan_code, r["code_text"])]

        for r in range(2, out_ws.max_row + 1):
            plan_name = out_ws.cell(r, name_col).value
            plan_code = out_ws.cell(r, code_col).value
            divisor_value = out_ws.cell(r, divisor_col).value if divisor_col else None
            divisor = as_decimal(divisor_value)

            if plan_name is None:
                continue
            if plan_code is not None and not isinstance(plan_code, int):
                try:
                    plan_code = int(str(plan_code))
                except Exception:
                    plan_code = None

            matches = find_matches(str(plan_name), plan_code)
            if not matches:
                continue

            a2_rows = [m for m in matches if m["degree"] == "A/2"]
            c3_rows = [m for m in matches if m["degree"] == "C/3"]

            for label_key, cols in metric_cols.items():
                metric_key = metric if label_key == "__default__" else label_key
                summary = summarize_metrics(a2_rows, c3_rows, metric_key)
                tonality_key = None if label_key == "__default__" else label_key
                write_summary(
                    out_ws,
                    r,
                    cols,
                    summary,
                    tonality_cols.get(tonality_key) if tonality_key else None,
                    divisor,
                )

        out_wb.save(output_path)
    finally:
        out_wb.close()
    return output_path


def main() -> int:
    args = parse_args()
    output_path = process_files(
        input_path=args.input,
        template_path=args.template,
        output_path=args.output,
        metric=args.metric,
        sheet=args.sheet,
        in_place=args.in_place,
    )
    print(f"Wrote: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
