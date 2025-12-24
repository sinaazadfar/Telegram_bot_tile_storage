from pathlib import Path

import openpyxl

from .config import TEMPLATE_LOCK, TEMPLATE_PATH
from .utils import clean_text, normalize_code_value, normalize_query


def resolve_template_path(template_path: Path | None) -> Path:
    if template_path is None:
        return TEMPLATE_PATH
    return template_path


def append_template_row(
    code: str, name: str, size: str, divisor, template_path: Path | None = None
) -> int:
    template_path = resolve_template_path(template_path)
    if not template_path.exists():
        raise FileNotFoundError("Template not found.")
    with TEMPLATE_LOCK:
        wb = openpyxl.load_workbook(template_path)
        try:
            ws = wb.active
            last_row = 1
            for r in range(2, ws.max_row + 1):
                if any(ws.cell(r, c).value not in (None, "") for c in range(1, 5)):
                    last_row = r
            next_row = last_row + 1
            ws.cell(next_row, 1).value = code
            ws.cell(next_row, 2).value = name
            ws.cell(next_row, 3).value = size
            ws.cell(next_row, 4).value = divisor
            wb.save(template_path)
        finally:
            wb.close()
    return next_row


def list_template_rows(template_path: Path | None = None) -> list[dict]:
    template_path = resolve_template_path(template_path)
    if not template_path.exists():
        raise FileNotFoundError("Template not found.")
    rows: list[dict] = []
    with TEMPLATE_LOCK:
        wb = openpyxl.load_workbook(template_path)
        try:
            ws = wb.active
            for r in range(2, ws.max_row + 1):
                code_raw = ws.cell(r, 1).value
                name_raw = ws.cell(r, 2).value
                size_raw = ws.cell(r, 3).value
                divisor_raw = ws.cell(r, 4).value
                code_display = normalize_code_value(code_raw)
                name_display = clean_text(name_raw)
                size_display = clean_text(size_raw)
                divisor_display = normalize_code_value(divisor_raw)
                if not code_display and not name_display:
                    continue
                rows.append(
                    {
                        "row": r,
                        "code_raw": code_raw,
                        "name_raw": name_raw,
                        "size_raw": size_raw,
                        "divisor_raw": divisor_raw,
                        "code_display": code_display,
                        "name_display": name_display,
                        "size_display": size_display,
                        "divisor_display": divisor_display,
                    }
                )
        finally:
            wb.close()
    return rows


def find_template_matches_any(query: str, template_path: Path | None = None) -> list[dict]:
    rows = list_template_rows(template_path)
    query_norm = normalize_query(query)
    if not query_norm:
        return rows
    matches: list[dict] = []
    for row in rows:
        code_norm = normalize_query(row["code_display"])
        name_norm = normalize_query(row["name_display"])
        if query_norm in code_norm or query_norm in name_norm:
            matches.append(row)
    return matches


def delete_template_row(row: dict, template_path: Path | None = None) -> bool:
    template_path = resolve_template_path(template_path)
    if not template_path.exists():
        raise FileNotFoundError("Template not found.")
    target_row = row.get("row")
    target_code_norm = normalize_code_value(
        row.get("code_raw", row.get("code_display", ""))
    )
    target_name_norm = normalize_query(
        row.get("name_raw", row.get("name_display", ""))
    )
    with TEMPLATE_LOCK:
        wb = openpyxl.load_workbook(template_path)
        try:
            ws = wb.active
            if isinstance(target_row, int) and 2 <= target_row <= ws.max_row:
                row_code = normalize_code_value(ws.cell(target_row, 1).value)
                row_name = normalize_query(ws.cell(target_row, 2).value)
                if row_code == target_code_norm and row_name == target_name_norm:
                    ws.delete_rows(target_row, 1)
                    wb.save(template_path)
                    return True
            for r in range(2, ws.max_row + 1):
                row_code = normalize_code_value(ws.cell(r, 1).value)
                row_name = normalize_query(ws.cell(r, 2).value)
                if row_code == target_code_norm and row_name == target_name_norm:
                    ws.delete_rows(r, 1)
                    wb.save(template_path)
                    return True
        finally:
            wb.close()
    return False


def update_template_row(
    original: dict, new_values: dict, template_path: Path | None = None
) -> bool:
    template_path = resolve_template_path(template_path)
    if not template_path.exists():
        raise FileNotFoundError("Template not found.")
    target_row = original.get("row")
    target_code_norm = normalize_code_value(
        original.get("code_raw", original.get("code_display", ""))
    )
    target_name_norm = normalize_query(
        original.get("name_raw", original.get("name_display", ""))
    )
    with TEMPLATE_LOCK:
        wb = openpyxl.load_workbook(template_path)
        try:
            ws = wb.active
            row_index = None
            if isinstance(target_row, int) and 2 <= target_row <= ws.max_row:
                row_code = normalize_code_value(ws.cell(target_row, 1).value)
                row_name = normalize_query(ws.cell(target_row, 2).value)
                if row_code == target_code_norm and row_name == target_name_norm:
                    row_index = target_row
            if row_index is None:
                for r in range(2, ws.max_row + 1):
                    row_code = normalize_code_value(ws.cell(r, 1).value)
                    row_name = normalize_query(ws.cell(r, 2).value)
                    if row_code == target_code_norm and row_name == target_name_norm:
                        row_index = r
                        break
            if row_index is None:
                return False
            ws.cell(row_index, 1).value = new_values["code"]
            ws.cell(row_index, 2).value = new_values["name"]
            ws.cell(row_index, 3).value = new_values["size"]
            ws.cell(row_index, 4).value = new_values["divisor"]
            wb.save(template_path)
            return True
        finally:
            wb.close()


def get_output_row_details(target: dict, output_path) -> list[tuple[str, str]]:
    if not output_path.exists():
        raise FileNotFoundError("Output not found.")
    target_code_norm = normalize_code_value(
        target.get("code_raw", target.get("code_display", ""))
    )
    target_name_norm = normalize_query(
        target.get("name_raw", target.get("name_display", ""))
    )
    wb = openpyxl.load_workbook(output_path, data_only=True)
    try:
        ws = wb.active
        code_col = 1
        name_col = 2
        row_index = None
        for r in range(2, ws.max_row + 1):
            row_code = normalize_code_value(ws.cell(r, code_col).value)
            row_name = normalize_query(ws.cell(r, name_col).value)
            if target_code_norm and row_code != target_code_norm:
                continue
            if target_name_norm and row_name != target_name_norm:
                continue
            row_index = r
            break
        if row_index is None:
            return []
        details: list[tuple[str, str]] = []
        for col in range(1, ws.max_column + 1):
            header = clean_text(ws.cell(1, col).value)
            if not header:
                continue
            value = ws.cell(row_index, col).value
            if value in (None, ""):
                continue
            details.append((header, str(value)))
        return details
    finally:
        wb.close()
